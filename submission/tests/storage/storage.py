import datetime as date
from datetime import datetime
import openpyxl
messageList = []
deletedMessages = 0  
def checkBooks():
    try:
        wb1 = openpyxl.load_workbook('example.xlsx')
        ws1 = wb1.get_sheet_by_name('Sheet1')
    except:
        filepath = "example.xlsx"
        wb = openpyxl.Workbook()
        wb.create_sheet("Sheet1")
        wb.save(filepath)
    try:
        wb2 = openpyxl.load_workbook('example1.xlsx')
        ws2 = wb2.get_sheet_by_name('Sheet1')
    except:
        filepath = "example1.xlsx"
        wb2 = openpyxl.Workbook()
        wb2.create_sheet("Sheet1")
        wb2.save(filepath)
    try:
        wb1 = openpyxl.load_workbook('example.xlsx')
        ws1 = wb1.get_sheet_by_name('Sheet1')
        date_object_test = datetime.strptime(str(ws1["E1"]), '%Y-%m-%d')
    except:
        ws1["A1"] = "DO"
        ws1["B1"] = "NOT"
        ws1["C1"] = "DELETE"
        ws1["D1"] = "ME"
        ws1["E1"] = "2019-11-24"
        ws1["F1"] = "00:00"
        wb1.save('example.xlsx')
    
def write(message):
    checkBooks()
    #print(message + "Type: " + type(message))
    wb1 = openpyxl.load_workbook('example.xlsx')
    ws1 = wb1.get_sheet_by_name('Sheet1')
    ws1.append(message)
    wb1.save('example.xlsx')
    
def createFakeMessage(month, x):
    checkBooks()
    d = {}
    d[0] = "example message text"
    d[1] = "jeffB@dogefin.jp"
    d[2] = "jeffA@dogefin.jp"
    d[3] = x
    months = (date.date.today() - date.timedelta(31 * month))
    d[4] = months.strftime('%Y-%m-%d')
    now = datetime.today()
    d[5] = now.strftime("%H:%M")
    write(list(d.values()))

def read():
    checkBooks()
    wb1 = openpyxl.load_workbook('example.xlsx')
    ws1 = wb1.get_sheet_by_name('Sheet1')
    
    wb2 = openpyxl.load_workbook('example1.xlsx')
    ws2 = wb2.get_sheet_by_name('Sheet1')
    wb2.remove_sheet(ws2)
    wb2.create_sheet("Sheet1")
    ws2 = wb2.get_sheet_by_name('Sheet1')
    
    rowNumbers = [] #append to this to store relavent row numbers
    for cell in ws1['E']:
        global deletedMessages
        date_object = datetime.strptime(str(cell.value), '%Y-%m-%d')
        sixmonths = (date.date.today() - date.timedelta(182))
        if date_object.date() <= sixmonths and ws1['D' + str(cell.row)].value == "N":
            deletedMessages +=1
            print(date_object.date(), " was before ", sixmonths)
        else:
            #print(ws1['B' + str(cell.row)].value) #incredible
            print(date_object.date(), " was after or important", sixmonths)
            rowNumbers.append(cell.row)
            
    for i in rowNumbers:
        global messageList
        l = []
        l.append(ws1["A" + str(i)].value)
        l.append(ws1["B" + str(i)].value)
        l.append(ws1["C" + str(i)].value)
        l.append(ws1["D" + str(i)].value)
        l.append(ws1["E" + str(i)].value)
        l.append(ws1["F" + str(i)].value)
        d = {}
        d[0] = ws1["A" + str(i)].value
        d[1] = ws1["B" + str(i)].value
        d[2] = ws1["C" + str(i)].value
        d[3] = ws1["D" + str(i)].value
        d[4] = ws1["E" + str(i)].value
        d[5] = ws1["F" + str(i)].value
        messageList.append(d)
        ws2.append(l)
        wb2.save('example.xlsx')
        
    print (str(deletedMessages) + " Messages Were Deleted")


#for x in range(12):
#    createFakeMessage(x, "Y")
#    createFakeMessage(x, "N")
    
a = str(input("enter 1 to write fake messages or 2 to delete them: "))
if a == "1":
    for x in range(12):
        createFakeMessage(x, "Y")
        createFakeMessage(x, "N")
elif a == "2":
    read()
    print(messageList)

else:
    print("Incorrect Input please enter a 1 or 2")
