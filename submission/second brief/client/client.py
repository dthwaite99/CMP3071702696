from PyQt5.QtCore import QDateTime, Qt, QTimer
from PyQt5.QtWidgets import (QApplication, QCheckBox, QComboBox, QDateTimeEdit,
        QDial, QDialog, QGridLayout, QGroupBox, QHBoxLayout, QLabel, QLineEdit,
        QProgressBar, QPushButton, QRadioButton, QScrollBar, QSizePolicy,
        QSlider, QSpinBox, QStyleFactory, QTableWidget, QTabWidget, QTextEdit,
        QVBoxLayout, QWidget, QMessageBox)
from socket import AF_INET, socket, SOCK_STREAM
from threading import Thread
from simplecrypt import encrypt, decrypt
import pickle
import openpyxl
import time
import sys
import datetime as date
from datetime import datetime
import winsound
import copy

HEADERSIZE = 10 #headersize for transmissions
password = 'password' # decryption password
unreadMessList = [] #list of all messages used in the changeView function

# All of the code for the GUI
app = QApplication([])

text_area = QTextEdit()
text_area.setFocusPolicy(Qt.NoFocus)

unread_mess = QTextEdit()
unread_mess.setFocusPolicy(Qt.NoFocus)

message = QLineEdit()

styleComboBox = QComboBox()

button = QPushButton("Important")
button.setCheckable(True)

layout = QVBoxLayout()
layout.addWidget(styleComboBox)
layout.addWidget(button)
layout.addWidget(text_area)
layout.addWidget(unread_mess)
layout.addWidget(message)

window = QWidget()
window.setLayout(layout)
window.show()
workbook_name = "messages.xlsx" #name of the worksheet I left this as a variable incase doge financial would prefer another name for it
def makeNoise(): # this function produces four notes and is used to alert the user of important messages
    winsound.Beep(2000, 500)
    winsound.Beep(2500, 500)
    winsound.Beep(2000, 500)
    winsound.Beep(2500, 500)
    
#Event handlers:
def checkBooks(): # this function makes sure all the workbooks exist and are ready to be written into and read from
    try:
        wb1 = openpyxl.load_workbook(workbook_name)
        ws1 = wb1.get_sheet_by_name('Sheet1')
    except:
        filepath = workbook_name
        wb = openpyxl.Workbook()
        wb.create_sheet("Sheet1")
        wb.save(filepath)
    
    try:
        wb2 = openpyxl.load_workbook('example1.xlsx')
        ws2 = wb2.get_sheet_by_name('Sheet1')
    except:
        filepath = "example1.xlsx"
        wb2 = openpyxl.Workbook()
        wb2.create_sheet("Sheet1")
        wb2.save(filepath)
    try:
        date_object_test = datetime.strptime(str(ws1["E1"]), '%Y-%m-%d') #if this fails it means there are no entrys and this will break the read function
    except: #writes a fake entry into the sheet so that read doesn't crash 
        ws1["A1"] = "DO"
        ws1["B1"] = "NOT"
        ws1["C1"] = "DELETE ME"
        ws1["D1"] = "Y"
        ws1["E1"] = "2019-11-24"
        ws1["F1"] = "00:00"
        wb1.save(workbook_name)
       
def write(message): # writes a message into the database
    checkBooks()
    #print(message + "Type: " + type(message))
    wb1 = openpyxl.load_workbook(workbook_name)
    ws1 = wb1.get_sheet_by_name('Sheet1')
    ws1.append(message)
    wb1.save(workbook_name)
    
def read(): # this function both deletes old messages and reads the ones that haven't been deleted into a list
    checkBooks()
    wb1 = openpyxl.load_workbook(workbook_name)
    ws1 = wb1.get_sheet_by_name('Sheet1')
    
    wb2 = openpyxl.load_workbook('example1.xlsx') #this workbook is used to store all of the messages worth keeping and then then it is saved as workbook_name which in effect deletes all unwanted messages
    ws2 = wb2.get_sheet_by_name('Sheet1')
    wb2.remove_sheet(ws2)
    wb2.create_sheet("Sheet1")
    ws2 = wb2.get_sheet_by_name('Sheet1')
    
    rowNumbers = [] #append to this to store relavent row numbers
    for cell in ws1['E']: # iterates through the column containing the date
        global deletedMessages
        date_object = datetime.strptime(str(cell.value), '%Y-%m-%d') #turns cell value into a datetime object
        sixmonths = (date.date.today() - date.timedelta(182)) # a datetime object representing siz months in the past from today
        if date_object.date() <= sixmonths and ws1['D' + str(cell.row)].value == "N": #tests if this message is older than six months and also unimportant
            deletedMessages +=1
            print(date_object.date(), " was before ", sixmonths)
        else:
            #print(ws1['B' + str(cell.row)].value) #incredible
            print(date_object.date(), " was after or important", sixmonths)
            rowNumbers.append(cell.row) # by making a list of rownumbers to be saved you can iterate through the list and recover the rows worth keeping
            
            
    for i in rowNumbers: #saves the row into the file and also the message list
        global unreadMessList
        l = []
        l.append(ws1["A" + str(i)].value)
        l.append(ws1["B" + str(i)].value)
        l.append(ws1["C" + str(i)].value)
        l.append(ws1["D" + str(i)].value)
        l.append(ws1["E" + str(i)].value)
        l.append(ws1["F" + str(i)].value)
        d = {}
        d[0] = ws1["A" + str(i)].value
        d[1] = ws1["B" + str(i)].value
        d[2] = ws1["C" + str(i)].value
        d[3] = ws1["D" + str(i)].value
        d[4] = ws1["E" + str(i)].value
        d[5] = ws1["F" + str(i)].value
        unreadMessList.append(d)
        
        #print(l)
        ws2.append(l)
        wb2.save(workbook_name) # saves workbook as the workbook name to ensure that only the correct files are kept


def send_message(): #function for sending message
    global activeChat #active chat is who the message is going to
    global user #user is where the message is coming from
    d = {} #messages are sent as pickled dictionaries
    d[0] = message.text() #message text
    d[1] = activeChat #recipient
    d[2] = user #sender
    if button.isChecked(): #checkkks whether or not message is important is makrs message accordingly
        d[3] = "Y" #y for yes
        text_area.append(str(date.date.today().strftime('%Y-%m-%d') + " You :" + message.text()+" IMPORTANT " + datetime.today().strftime("%H:%M")))
    else:
        d[3] = "N" # n for no
        text_area.append(str(date.date.today().strftime('%Y-%m-%d') + " You :" + message.text()+" NOT IMPORTANT " + datetime.today().strftime("%H:%M")))
    d[4] = date.date.today().strftime('%Y-%m-%d') #adds the date fo the message
    now = datetime.today() #current time
    d[5] = now.strftime("%H:%M") # adds the time to the message

    writeDic = copy.deepcopy(d) #write unencrypted message to the list and file for speed
    unreadMessList.append(writeDic)
    write(list(d.values()))
    d[0] = encrypt(password, message.text()) #encrypt message text
    msg = pickle.dumps(d) #pickle the dictionary
    msg = bytes(f"{len(msg):<{HEADERSIZE}}", 'utf-8')+msg #make message into bytes
    client_socket.send(msg) #send message
    message.clear() #clear the chatbox

def receive():
    global user #user
    global unreadMessList #message list
    full_msg = b'' #variables for LAN
    new_msg = True
    while True:
        msg = client_socket.recv(BUFSIZ) #waiting to recieve a transmission
        if new_msg:
            msglen = int(msg[:HEADERSIZE])
            new_msg = False

        full_msg += msg

        #print(len(full_msg))

        if len(full_msg)-HEADERSIZE == msglen:
            x = pickle.loads(full_msg[HEADERSIZE:])
            if x[1] == user: #if recipient is user then decypt and display message if not then do not process it
                display = decrypt(password, x[0]) #decrypts message
                if x[2] == activeChat: # if message sender is the active chat
                    if x[3] == "Y": #display code for if message is important
                        text_area.append(str(x[4] + ": Them :" + display.decode("utf-8")+" IMPORTANT " + x[5]))
                        makeNoise()

                    else: #if message unimporant then display message using this code
                        text_area.append(str(x[4] + ": Them :" + display.decode("utf-8")+" NOT IMPORTANT " + x[5]))
                    
                else:
                    if x[3] == "Y":
                        unread_mess.append("You missed an important message from " + x[2] + " at " + x[5])
                        makeNoise()
                    else:
                        unread_mess.append("You missed a message from " + x[2] + " at " + x[5])
                x[0] = display.decode("utf-8") #writes decoded message into the file and message list
                write(list(x.values()))
                unreadMessList.append(x)
        new_msg = True
        full_msg = b""
          
             
def checkUnreadMessage(name): #iterates through the message list and displays appropriate messages
    global unreadMessList
    for x in unreadMessList:
        display = x[0] 
        #print(type(display)) was having some issues with display unexpectidly being bytes
        if x[2] == name: #if the sender matches the name variable then display
            if x[3] == "Y":
                text_area.append(str(x[4] + ": Them :" + display + " IMPORTANT " + x[5]))
            else:
                        text_area.append(str(x[4] + ": Them :" + display + " NOT IMPORTANT " + x[5]))
        if x[2] == user and x[1] == name: #if the user sent a message to the name variable then display
            if x[3] == "Y":
                        text_area.append(str(x[4] + ": You :" + str(display) + " IMPORTANT " + x[5]))
            else:
                        text_area.append(str(x[4] + ": You :" + str(display) + " NOT IMPORTANT " + x[5]))
        #else:
            #print("test")
        
        
def changeView(): #tells you who yoiu are chatting with and displays all previous messages
    global activeChat
    text_area.clear() #clears chatbox
    text_area.append("You are now chatting with " + str(nameArray[styleComboBox.currentIndex()])) 
    activeChat = nameArray[styleComboBox.currentIndex()]
    checkUnreadMessage(activeChat) #calls function with the activechat as an argument
    
            
# Signals:
message.returnPressed.connect(send_message) #when you hit enter message is sent
styleComboBox.activated.connect(changeView) #when you change the dropdown box the chat gets updated

HOST = "127.0.0.1" #IP of server
PORT = 33000 #Port used by server

nameArray = [] #array of all contacts used for sign in and chats
try:
    wb1 = openpyxl.load_workbook('contacts.xlsx')
    ws1 = wb1.get_sheet_by_name('Sheet1')
    for cell in ws1['A']:
        nameArray.append(str(cell.value)) #loads names from contacts sheet
    print("Contacts loaded properly")
except:
    print("Your contacts File is Missing So We have Provided you with a defualt one Please Contact an Administrator")
    nameArray.append("jeffA@dogefin.jp")
    nameArray.append("jeffB@dogefin.jp")
    nameArray.append("jeffC@dogefin.jp")

user = input('Enter username: ') #only accepts users if they enter a username from the text sheet
if user in nameArray: #if entered username dosn't exist then reject log in
    print("You have logged in as '" + user + "' the program will now launch")
    nameArray.remove(user)
else:
    print("No match the username you entered either doesn't exist or you need an updated contacts file, please contact an administrator")
    time.sleep(5)
    sys.exit()

read()

styleComboBox.addItems(nameArray) #adds all names bar the users to the chat selector so that they cannot message themselves
activeChat = nameArray[0]
if not PORT:
    PORT = 33000
else:
    PORT = int(PORT)

BUFSIZ = 1024 #buffersize
ADDR = (HOST, PORT) #address of server

client_socket = socket(AF_INET, SOCK_STREAM) #creates socket
client_socket.connect(ADDR) #binds address to socket

receive_thread = Thread(target=receive) #starts thread so that it can always recieve messages no matter what
receive_thread.start() #starts thread

#time.sleep(5)
app.exec_() #starts GUI